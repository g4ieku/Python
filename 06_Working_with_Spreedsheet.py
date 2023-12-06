# working with spreadsheets
# package = openpyxl
import openpyxl
import math

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file['Sheet1']

# how many products per supplier
products_per_supplier = {}              #dictionary
total_value_per_supplier = {}
products_under_10_inv = {}

# print(product_list.max_row)

product_list.cell(1,5).value = 'Total Price'

for product_row in range(2, product_list.max_row+1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    total_price = inventory * price
    product_num = product_list.cell(product_row,1).value
    # print(supplier_name)
    
    
    # calculation number of products per supplier
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] +=1
    else:
        products_per_supplier[supplier_name] = 1
        
        
    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] = total_value_per_supplier[supplier_name] + total_price
    else:
        total_value_per_supplier[supplier_name] = total_price
        
        
    # logic products with inventory less than 10
    if inventory < 10:
        products_under_10_inv[int(product_num)] = int(inventory)
    
    
    product_list.cell(product_row,5).value = total_price  #and after this opperation we need to save the faile
    
inv_file.save('inventory_with_total_value.xlsx')
print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)